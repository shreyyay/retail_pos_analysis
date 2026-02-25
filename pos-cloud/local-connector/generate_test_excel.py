"""
Generate tally_test_data.xlsx — sample sales and purchase vouchers for Tally testing.
Run: python generate_test_excel.py
Then import the generated file into Tally manually or via XML.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = Workbook()

# ── Helpers ────────────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")
TOTAL_FONT = Font(bold=True, size=10)
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center")

thin = Side(style="thin", color="B0B0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

today = date.today()

def style_header(ws, headers, col_widths):
    ws.row_dimensions[1].height = 22
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

def style_row(ws, row, n_cols, alt=False):
    ws.row_dimensions[row].height = 18
    fill = ALT_FILL if alt else None
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.border    = BORDER
        cell.alignment = CENTER if col != 2 else LEFT   # Party/Item name left-aligned

def d(offset):
    return (today - timedelta(days=offset)).strftime("%d-%m-%Y")


# ── Sheet 1: Sales Vouchers ───────────────────────────────────────────────────

ws_sales = wb.active
ws_sales.title = "Sales Vouchers"
ws_sales.freeze_panes = "A2"

s_headers = ["Date", "Voucher No", "Party Name", "Item Name",
             "Qty", "Unit", "Rate (₹)", "Amount (₹)",
             "GST %", "CGST (₹)", "SGST (₹)", "Total (₹)"]
s_widths  = [13, 13, 18, 20, 8, 8, 12, 14, 8, 12, 12, 14]
style_header(ws_sales, s_headers, s_widths)

# 15 realistic Indian kirana/retail sales entries
sales_data = [
    # (days_ago, voucher_no, party,       item,              qty, unit,  rate,   gst)
    (0,  "S-001", "Cash",          "Basmati Rice 5kg",    10, "Bag",  550,   5),
    (0,  "S-002", "Ramesh Kumar",  "Toor Dal 1kg",        20, "Pkt",   95,   5),
    (0,  "S-003", "Cash",          "Sunflower Oil 1L",    15, "Btl",  130,   5),
    (1,  "S-004", "Suresh Traders","Sugar 1kg",           30, "Pkt",   45,   5),
    (1,  "S-005", "Cash",          "Wheat Flour 10kg",     8, "Bag",  380,   5),
    (1,  "S-006", "Meena Store",   "Tata Tea Premium",    12, "Pkt",  165,  12),
    (2,  "S-007", "Cash",          "Surf Excel 1kg",       6, "Pkt",  195,  18),
    (2,  "S-008", "Priya General", "Colgate Toothpaste",  10, "Pcs",   95,  18),
    (2,  "S-009", "Cash",          "Parle-G Biscuit",     24, "Pkt",   10,  18),
    (3,  "S-010", "Anand Kirana",  "Amul Butter 500g",     5, "Pcs",  280,  12),
    (3,  "S-011", "Cash",          "Good Day Biscuit",    18, "Pkt",   30,  18),
    (4,  "S-012", "Cash",          "Vim Dish Wash Bar",   10, "Pcs",   30,  18),
    (4,  "S-013", "Ravi Stores",   "Lifebuoy Soap",       20, "Pcs",   35,  18),
    (5,  "S-014", "Cash",          "Maggi Noodles 70g",   30, "Pkt",   14,  18),
    (5,  "S-015", "Lakshmi Mart",  "Fortune Soyabean Oil", 10,"Btl",  115,   5),
]

for i, (days, vno, party, item, qty, unit, rate, gst) in enumerate(sales_data, 2):
    amount = qty * rate
    cgst   = round(amount * gst / 200, 2)
    sgst   = round(amount * gst / 200, 2)
    total  = amount + cgst + sgst
    row = [d(days), vno, party, item, qty, unit, rate, amount, f"{gst}%", cgst, sgst, total]
    for col, val in enumerate(row, 1):
        ws_sales.cell(row=i, column=col, value=val)
    style_row(ws_sales, i, len(s_headers), alt=(i % 2 == 0))

# Total row
total_row = len(sales_data) + 2
ws_sales.cell(row=total_row, column=7, value="TOTAL").font = TOTAL_FONT
ws_sales.cell(row=total_row, column=8,
              value=f"=SUM(H2:H{total_row-1})").font = TOTAL_FONT
ws_sales.cell(row=total_row, column=10,
              value=f"=SUM(J2:J{total_row-1})").font = TOTAL_FONT
ws_sales.cell(row=total_row, column=11,
              value=f"=SUM(K2:K{total_row-1})").font = TOTAL_FONT
ws_sales.cell(row=total_row, column=12,
              value=f"=SUM(L2:L{total_row-1})").font = TOTAL_FONT


# ── Sheet 2: Purchase Vouchers ────────────────────────────────────────────────

ws_pur = wb.create_sheet("Purchase Vouchers")
ws_pur.freeze_panes = "A2"

p_headers = ["Date", "Voucher No", "Supplier Name", "Item Name",
             "Qty", "Unit", "Rate (₹)", "Amount (₹)", "GST %",
             "CGST (₹)", "SGST (₹)", "Total (₹)"]
p_widths  = [13, 13, 20, 20, 8, 8, 12, 14, 8, 12, 12, 14]
style_header(ws_pur, p_headers, p_widths)

purchase_data = [
    (2,  "P-001", "Agro Wholesale",    "Basmati Rice 5kg",     50, "Bag",  490,  5),
    (3,  "P-002", "National Traders",  "Toor Dal 1kg",        100, "Pkt",   80,  5),
    (3,  "P-003", "Sunfeast Distributor","Wheat Flour 10kg",   40, "Bag",  330,  5),
    (4,  "P-004", "HUL Distributor",   "Surf Excel 1kg",       30, "Pkt",  160, 18),
    (5,  "P-005", "ITC Limited",       "Tata Tea Premium",     60, "Pkt",  140, 12),
    (6,  "P-006", "Parle Distributor", "Parle-G Biscuit",     120, "Pkt",    8, 18),
    (6,  "P-007", "Amul Agency",       "Amul Butter 500g",     20, "Pcs",  240, 12),
    (7,  "P-008", "Ruchi Soya",        "Sunflower Oil 1L",     60, "Btl",  105,  5),
]

for i, (days, vno, supplier, item, qty, unit, rate, gst) in enumerate(purchase_data, 2):
    amount = qty * rate
    cgst   = round(amount * gst / 200, 2)
    sgst   = round(amount * gst / 200, 2)
    total  = amount + cgst + sgst
    row = [d(days), vno, supplier, item, qty, unit, rate, amount, f"{gst}%", cgst, sgst, total]
    for col, val in enumerate(row, 1):
        ws_pur.cell(row=i, column=col, value=val)
    style_row(ws_pur, i, len(p_headers), alt=(i % 2 == 0))

total_row = len(purchase_data) + 2
ws_pur.cell(row=total_row, column=7, value="TOTAL").font = TOTAL_FONT
ws_pur.cell(row=total_row, column=8,  value=f"=SUM(H2:H{total_row-1})").font = TOTAL_FONT
ws_pur.cell(row=total_row, column=12, value=f"=SUM(L2:L{total_row-1})").font = TOTAL_FONT


# ── Sheet 3: How to Use ───────────────────────────────────────────────────────

ws_info = wb.create_sheet("How to Import in Tally")
info_lines = [
    ("HOW TO IMPORT THIS DATA INTO TALLY PRIME", True),
    ("", False),
    ("Method 1 — Manual Entry (Recommended for testing)", True),
    ("  1. Open Tally Prime with your company loaded", False),
    ("  2. Go to Vouchers → Sales", False),
    ("  3. Enter each row from the 'Sales Vouchers' sheet one by one", False),
    ("  4. Do the same for Purchase Vouchers", False),
    ("", False),
    ("Method 2 — XML Import (Faster, for bulk data)", True),
    ("  1. Run: python create_tally_xml.py  (generates tally_import.xml)", False),
    ("  2. In Tally Prime: Help → Import → Data", False),
    ("  3. Select tally_import.xml", False),
    ("  4. Tally will import all vouchers automatically", False),
    ("", False),
    ("NOTES", True),
    ("  - Date format in Tally: DD-MM-YYYY", False),
    ("  - 'Cash' party name works without creating a ledger (it's pre-built)", False),
    ("  - For other party names, create the ledger in Tally first:", False),
    ("      Accounts Info → Ledgers → Create → Group: Sundry Debtors", False),
    ("  - GST rates require GST configuration in Tally to be active", False),
    ("  - If GST is not configured, enter items without GST columns", False),
]
ws_info.column_dimensions["A"].width = 70
for r, (text, bold) in enumerate(info_lines, 1):
    cell = ws_info.cell(row=r, column=1, value=text)
    cell.font = Font(bold=bold, size=11 if bold else 10,
                     color="1F4E79" if bold else "000000")
    cell.alignment = LEFT
    ws_info.row_dimensions[r].height = 20


# ── Save ──────────────────────────────────────────────────────────────────────

out = "tally_test_data.xlsx"
wb.save(out)
print(f"Created: {out}")
print(f"  Sales Vouchers:    {len(sales_data)} entries")
print(f"  Purchase Vouchers: {len(purchase_data)} entries")
